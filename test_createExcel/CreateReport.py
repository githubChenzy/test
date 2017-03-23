
#-*- coding:utf-8 -*-

import os
import time
import logging
import logging.config
from collections import OrderedDict
from openpyxl import Workbook

import setting
from db import db_mysql

logging.config.dictConfig(setting.LOGGING)
logger = logging.getLogger('mylogger')


class RegisterList(object):

	decimal_fields = ['amount', 'real_amount', 'real_ptb',]

	def __init__(self):
		self.titleHolderLine = 1
		self.pageSize = 50
		self.wb = Workbook()
		self.sheetName = self.setSheetname()

		self.fieldNameMap = self.setFieldNameMap()
		self.fieldNameMap.update(self.joinQueryFieldNameMap())
		print('fieldNameMap-->', self.fieldNameMap)

		self.initializeSheet()

	def setSheetname(self):
		return u'用户注册表'

	def setFieldNameMap(self):
		return {
			'username': u'注册账号',
			'flag': u'账号状态',
			'reg_time': u'注册时间',
			'imeil': u'IMEI码',
			'agent': u'渠道名称',   # 使用 cy_department表中去查询顶级渠道名
			'channel_id': u'顶级渠道',
			'gameid': u'游戏'  # cy_game table
		}

	def joinQueryFieldNameMap(self):
		return OrderedDict({
			# 'status': u'支付状态',  # cy_pay table
			'login_time': u'最后一次登录时间', # cy_logininffo
			'count(id)': u'累计充值次数', # cy_pay table
			'sum(amount)': u'累计充值金额' # cy_pay table
		})

	# self.dbfield 为一个元组 数据库查找返回的字段名
	# username flag reg_time imeil gameid agent channel_id
	def createPage(self, page, data):
		thisPageRowStartAt = (
			(page - 1) * self.pageSize + self.titleHolderLine + 1 # 每页的开始
		)
		id_index = self.dbfield.index('id') # 获得 id 字段的索引值，用于多表查询
		for row in range(0, len(data)): # 从数据库查询的数据条数，data为一个元组，每个也为一个元组
			availableColIdx = 0
			for col in range(0, len(self.dbfield)):  # 数据库的字段索引
				field_name = self.dbfield[col]		# 获得字段名
				if field_name in self.setFieldNameMap(): #
					availableColIdx += 1
					value = self._handle_excelvalue(self.dbfield[col], data[row][col])
				elif field_name in self.joinQueryFieldNameMap():
					print('id value', data[row][id_index])
					availableColIdx += 1
					value = self._handle_excelvalue(self.dbfield[col], data[row][id_index])
					if field_name in self.decimal_fields:
						value = float(value);
					else:
						value = (u'%s' % value);
				self.ws.cell(
					row=thisPageRowStartAt + row,
					column=availableColIdx
				).value = value

	def _handle_excelvalue(self, filed, value):
		print('filed value', filed, value)

		if (value == u'None') or (value == u'default') or (value == None) or (value == 'DEFAULT') or (value == u'defalut'):
			return value
		pay_list_status_desc = {
			0: '待支付',
			1: '支付成功',
			2: '失败',
		}

		user_status = {
			0: '正常',
			1: '冻结'
		}
		if filed == "flag":
			value = user_status.get(value, "未知状态")
			print('flag: ', value)

		if filed == 'reg_time':
			value = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(float(value)))
			print('reg_time: ', value)
		# if filed == 'channel_id':
		# 	value = TopDepartmentCache().getTopDepartmentName(value);

		if filed == "count(id)":
			sql = 'select count(id) from cy_pay where userid=%s and status=1;' % value
			_, value = db_mysql.ExcuteSql().query_sql(sql)
			value = value[0]
			print("count id: ", value[0])

		if filed == 'sum(amount)':
			sql = 'select sum(amount) from cy_pay where userid=%s and status=1;'% value
			_, value = db_mysql.ExcuteSql().query_sql(sql)
			value = value[0]
			print('sum amount: ', value[0])

		if filed == 'login_time':
			sql = 'select max(login_time) from cy_logininfo where userid=%s;' % value
			_, value = db_mysql.ExcuteSql().query_sql(sql)
			print('login_time: ', value[0][0])
			value = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(value[0][0]))

		if filed == 'agent':
			print(value)
			try:
				if isinstance(int(value), int) or isinstance(int(value), long):
					sql = 'select name from cy_department where id=%s' % value
					_, value = db_mysql.ExcuteSql().query_sql(sql)
					value = value if value else '---'
			except Exception as e:
				logger.error(e)
				raise e

			print('agent: ', value)

		if filed == "gameid" :
			sql = 'select name from cy_game where id=%s' % value
			_, value = db_mysql.ExcuteSql().query_sql(sql)
			value = value[0][0] if len(value) else '---'
			print('before gameid: ', value)
			print('gameid: ', value)

		return value;

	def create(self):
		# sql = self.queueValue.get("sql");
		sql = 'select username, flag, reg_time, imeil, agent, channel_id, gameid, id from cy_members'
		page = 1;
		while page < 5:
			# print("准备获取第 %d 页数据, 从 %d 开始取 %d 条" % (page, (page - 1) * self.pageSize, self.pageSize));
			thisPageSql = (sql
				+ (
					" LIMIT %d, %d"
					% (
						(page - 1) * self.pageSize,
						self.pageSize
					)
				)
			)
			dbFiled, dbData = db_mysql.ExcuteSql().query_sql(thisPageSql);
			print("query data from mysql: ", dbFiled, dbData)

			if not dbData:
				break;
			if 1 == page:
				self.setTitle(dbFiled);
				logger.info(u"开始进行报表数据填充(任务略微有点耗时,请耐心等待)....")

			self.createPage(page, dbData);
			page += 1;

		return self.finish();

	def initializeSheet(self):
		self.ws = self.wb.create_sheet(
			self.sheetName,
			0
		)

	def setTitle(self, dbfield):
		self.dbfield = list(dbfield) + list(self.joinQueryFieldNameMap().keys())
		title_row = []
		print("settile dbfield", dbfield)
		print('self.fieldNameMap', self.fieldNameMap)
		for field in self.dbfield:
			if field in self.fieldNameMap:
				title_row.append(self.fieldNameMap.get(field))
		print('title_row', title_row)
		self.ws.append(title_row)

	@staticmethod
	def create_reportDir(path):
		if not os.path.isdir(path):
			os.mkdir(path)

	def finish(self):
		report_dir = os.path.join(os.path.dirname(__file__), 'report')
		self.create_reportDir(report_dir)
		filename = str(int(time.time())) + '.xlsx'
		self.wb.save(os.path.join(report_dir, filename))

if __name__ == "__main__":
	r = RegisterList()
	r.create()
	print("end")
